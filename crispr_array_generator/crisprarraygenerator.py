"""
Checks gRNAs for errors and generates CRISPR arrays from gRNAs 
Created: Fall 2023
Author: Willow Chernoske
"""


import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import crispr_array_generator.constants as cn


class Array(object):
    """
    Class containing all functions needed check gRNAs and generate CRISPR arrays
    """
    def extract_excel_data(excel_file):
        """
        Exracts all data from an excel file and isolates DNA
        components
        Args:
            string: name of excel file listing guide RNAs in 5' 
            to 3' format
        Returns:
            array: all DNA components listed in the input excel file
        """
        excel_file = excel_file+".xlsx"
        workbook = load_workbook(excel_file)
        exceldata = []
        grnas = []
        sheet = workbook.active
        # Changing the layout of each excel cell value
        for value in sheet.iter_rows(values_only=True):
            value = str(value)
            value = value.replace("'" , "")
            value = value.replace("," , "")
            value = value.replace("(" , "")
            value = value.replace(")" , "")
            exceldata.append(value)
        # Isolating DNA values
        for input in exceldata:
            valid_dna = all(i in cn.VALID_DNA for i in input)
            if valid_dna:
                grnas.append(input)
        return grnas

    def get_reverse_complement(dna):
        """
        Converts a 5' to 3' DNA string into its 3' to 5' reverse 
        complement
        Args:
            string: DNA to be converted
        Returns:
            string: reverse complement of input DNA
        """
        dna_comp = [cn.BASE_PAIRS[base] for base in dna]
        dna_rev_comp = dna_comp[::-1]
        return ''.join(dna_rev_comp)

    def make_columns_best_fit(excel_file):
        """
        Formats columns in excel files to the width of the text
        in each cell
        Args:
            string: name of excel file
        Returns:
            file: excel file with best fit column widths
        """
        excel_file = excel_file+".xlsx"
        workbook = load_workbook(excel_file)
        for sheet_name in workbook.sheetnames:
            for column_cells in workbook[sheet_name].columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = get_column_letter(column_cells[0].column)
                if new_column_length > 0:
                    workbook[sheet_name].column_dimensions[new_column_letter].width = new_column_length*1.23
        workbook.save(excel_file)

    def check_grna(grnas):
        """
        Takes gRNAs listed in an excel file and checks them for 
        self-targeting or length errors
        Args:
            string: name of excel file listing guide RNAs in 5' 
            to 3' format
        Returns:
            file: excel file listing processed gRNAs and errors 
            if found
        """
        if isinstance(grnas, str):
            grnas = Array.extract_excel_data(grnas)
        if isinstance(grnas, list):
            for grna in grnas:
                valid_dna = all(i in cn.VALID_DNA for i in grna)
            if not valid_dna:
                grnas.remove(grna)
        new_grnas = []
        row = 1
        # Creating the output workbook object
        excel_output = Workbook()
        sheet_1 = excel_output.create_sheet("gRNA check")
        if 'Sheet' in excel_output.sheetnames:
            excel_output.remove(excel_output['Sheet'])
        sheet_1.cell(row=1 , column=1).value = "gRNAs"
        sheet_1.cell(row=1 , column=3).value = "length error (>24 nucleotides)"
        sheet_1.cell(row=1 , column=4).value = "length error (<20 nucleotides)"
        for cell in sheet_1["1:1"]:
            cell.font = Font(bold = True)
        # Processing gRNAs and checking for errors
        for grna in grnas:
            row = row + 1
            #Removing CRISPR cut site within gRNA if present
            grna = grna.removeprefix("TTC")
            grna = grna.removeprefix("ttc")
            #Checking gRNA length
            if len(grna) > 24:
                sheet_1.cell(row=row , column=3).value = "X"
            if len(grna) < 20:
                c3 = sheet_1.cell(row=row , column=4).value = "X"
            new_grnas.append(grna)
            #Putting processed gRNAs into the output excel file
            c3 = sheet_1.cell(row=row , column=1)
            c3.value = grna
        excel_output.save('array_report.xlsx')
        Array.make_columns_best_fit('array_report')
        return new_grnas

    def get_array(grnas):
        """
        Generates ready-to-order oligos from input gRNAs listed on an 
        excel file
        Args:
            string: name of excel file listing guide RNAs in 5' to 3'
            format 
        Returns:
            excel file listing ready-to-order oligos for each set of 
            gRNAs and any errors if present
        """
        revcomp_grnas = []
        new_grnas = Array.check_grna(grnas)
        excel_output = load_workbook("array_report.xlsx")
        # Setting up the Array sheet
        sheet_2 = excel_output.create_sheet("Array")
        sheet_2.cell(row=1 , column=2).value = "Full array fwd:"
        sheet_2.cell(row=1 , column=3).value = "Full array rev:"
        sheet_2.cell(row=1 , column=5).value = "Errors:"
        sheet_2.cell(row=4 , column=1).value = "gRNA #:"
        sheet_2.cell(row=4 , column=2).value = "Fwd oligos:"
        sheet_2.cell(row=4 , column=3).value = "Rev oligos:"
        number = len(new_grnas)
        for value in range(1, number+1):
            sheet_2.cell(row=value + 4, column=1).value = value
        for cell in sheet_2["1:1"]:
            cell.font = Font(bold = True)
        for cell in sheet_2["4:4"]:
            cell.font = Font(bold = True)
        for cell in sheet_2["A:A"]:
            cell.font = Font(bold = True)
        # Checking for input error
        if number > 9:
            sheet_2.cell(row=2, column=5).value = "More than 9 gRNAs were identified. Please input 9 or fewer gRNAs per array."
        # Creating array oligos and inserting them into "array_report.xlsx"
        for grna in new_grnas:
            grnarev = Array.get_reverse_complement(grna)
            revcomp_grnas.append(grnarev)
        if number>=1 and number<=9:
            fwd_1 = sheet_2.cell(row=5 , column=2).value = "CCCTAAATAATTTCTACTGTTGTAGAT" + new_grnas[0]
            if number==1:
                rev_11 = sheet_2.cell(row=5 , column=3).value = "CGTT" + revcomp_grnas[0] + "ATCTACAACAGTAGAAATTATTT"
                sheet_2.cell(row=2 , column=2).value = fwd_1
                sheet_2.cell(row=2 , column=3).value = rev_11
        if number>=2 and number<=9:
            rev_1 = sheet_2.cell(row=5 , column=3).value = "GCCA" + revcomp_grnas[0] + "ATCTACAACAGTAGAAATTATTT"
            if number==2:
                fwd_22 = sheet_2.cell(row=6 , column=2).value = "TGGCAAATAATTTCTACTGTTGTAGAT" + new_grnas[1]
                rev_22 = sheet_2.cell(row=6 , column=3).value = "CGTT" + revcomp_grnas[1] + "ATCTACAACAGTAGAAATTATTT"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_22
                sheet_2.cell(row=2 , column=3).value = rev_22 + rev_1
        if number>=3 and number<=9:
            fwd_2 = sheet_2.cell(row=6 , column=2).value = "TGGCAAATAATTTCTACTGTTGTAGAT" + new_grnas[1] + "TTCT"
            rev_2 = sheet_2.cell(row=6 , column=3).value = revcomp_grnas[1] + "ATCTACAACAGTAGAAATTATTT"
            fwd_3 = sheet_2.cell(row=7 , column=2).value = "AAATAATTTCTACTGTTGTAGAT" + new_grnas[2]
            if number==3:
                rev_33 = sheet_2.cell(row=7 , column=3).value = "CGTT" + revcomp_grnas[2] + "ATCTACAACAGTAGAAATTATTTAGAA"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3
                sheet_2.cell(row=2 , column=3).value = rev_33 + rev_2 + rev_1
        if number>=4 and number<=9:
            rev_3 = sheet_2.cell(row=7 , column=3).value = "ATTG" + revcomp_grnas[2] + "ATCTACAACAGTAGAAATTATTTAGAA"
            if number==4:
                fwd_44 = sheet_2.cell(row=8 , column=2).value = "CAATAAATAATTTCTACTGTTGTAGAT" + new_grnas[3]
                rev_44 = sheet_2.cell(row=8 , column=3).value = "CGTT" + revcomp_grnas[3] + "ATCTACAACAGTAGAAATTATTT"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3 + fwd_44
                sheet_2.cell(row=2 , column=3).value = rev_44 + rev_3 + rev_2 + rev_1
        if number>=5 and number<=9:
            fwd_4 = sheet_2.cell(row=8 , column=2).value = "CAATAAATAATTTCTACTGTTGTAGAT" + new_grnas[3] + "TATG"
            rev_4 = sheet_2.cell(row=8 , column=3).value =  revcomp_grnas[3] + "ATCTACAACAGTAGAAATTATTT"
            fwd_5 = sheet_2.cell(row=9 , column=2).value = "AAATAATTTCTACTGTTGTAGAT" + new_grnas[4]
            if number==5:
                rev_55 = sheet_2.cell(row=9, column=3).value = "CGTT" + revcomp_grnas[4] + "ATCTACAACAGTAGAAATTATTTCATA"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3 + fwd_4 + fwd_5
                sheet_2.cell(row=2 , column=3).value = rev_55 + rev_4 + rev_3 + rev_2 + rev_1
        if number>=6 and number<=9:
            rev_5 = sheet_2.cell(row=9 , column=3).value = "TTCT" + revcomp_grnas[4] + "ATCTACAACAGTAGAAATTATTTCATA"
            if number==6:
                fwd_66 = sheet_2.cell(row=10 , column=2).value = "AGAAAAATAATTTCTACTGTTGTAGAT" + new_grnas[5]
                rev_66 = sheet_2.cell(row=10 , column=3).value = "CGTT" + revcomp_grnas[5] + "ATCTACAACAGTAGAAATTATTT"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3 + fwd_4 + fwd_5 + fwd_66
                sheet_2.cell(row=2 , column=3).value = rev_66 + rev_5 + rev_4 + rev_3 + rev_2 + rev_1
        if number>=7 and number<=9:
            fwd_6 = sheet_2.cell(row=10 , column=2).value = "AGAAAAATAATTTCTACTGTTGTAGAT" + new_grnas[5] + "TACA"
            rev_6 = sheet_2.cell(row=10 , column=3).value = revcomp_grnas[5] + "ATCTACAACAGTAGAAATTATTT"
            fwd_7 = sheet_2.cell(row=11 , column=2).value = "AAATAATTTCTACTGTTGTAGAT" + new_grnas[6]
            if number==7:
                rev_77 = sheet_2.cell(row=11 , column=3).value = "CGTT" + revcomp_grnas[6] + "ATCTACAACAGTAGAAATTATTTTGTA"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3 + fwd_4 + fwd_5 + fwd_6 + fwd_7
                sheet_2.cell(row=2 , column=3).value = rev_77 + rev_6 + rev_5 + rev_4 + rev_3 + rev_2 + rev_1
        if number>=8 and number<=9:
            rev_7 = sheet_2.cell(row=11, column=3).value = "CAGC" + revcomp_grnas[6] + "ATCTACAACAGTAGAAATTATTTTGTA"
            if number==8:
                fwd_88 = sheet_2.cell(row=12 , column=2).value = "GCTGAAATAATTTCTACTGTTGTAGAT" + new_grnas[7]
                rev_88 = sheet_2.cell(row=12 , column=3).value = "CGTT" + revcomp_grnas[7] + "ATCTACAACAGTAGAAATTATTT"
                sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3 + fwd_4 + fwd_5 + fwd_6 + fwd_7 + fwd_88
                sheet_2.cell(row=2 , column=3).value = rev_88 + rev_7 + rev_6 + rev_5 + rev_4 + rev_3 + rev_2 + rev_1
        if number==9:
            fwd_8 = sheet_2.cell(row=12 , column=2).value = "GCTGAAATAATTTCTACTGTTGTAGAT" + new_grnas[7] + "GAGT"
            rev_8 = sheet_2.cell(row=12 , column=3).value = revcomp_grnas[7] + "ATCTACAACAGTAGAAATTATTT"
            fwd_9 = sheet_2.cell(row=13 , column=2).value = "AAATAATTTCTACTGTTGTAGAT" + new_grnas[8]
            rev_9 = sheet_2.cell(row=13 , column=2).value = "CGTT" + revcomp_grnas[8] + "ATCTACAACAGTAGAAATTATTTACTC"
            sheet_2.cell(row=2 , column=2).value = fwd_1 + fwd_2 + fwd_3 + fwd_4 + fwd_5 + fwd_6 + fwd_7 + fwd_8 + fwd_9
            sheet_2.cell(row=2 , column=3).value = rev_9 + rev_8 + rev_7 + rev_6 + rev_5 + rev_4 + rev_3 + rev_2 + rev_1
        excel_output.save('array_report.xlsx')
        Array.make_columns_best_fit('array_report')


