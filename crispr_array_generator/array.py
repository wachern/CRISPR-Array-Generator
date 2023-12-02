import crispr_array_generator.constants as cn
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


class Array(object):
    def extract_excel_data(excel_file):
        """
        Exracts all data from an excel file and isolates DNA components
        Args:
        string: name of excel file listing guide RNAs in 5' to 3' format
        Returns:
        array: all DNA components listed in the input excel file
        """
        
        excel_file = excel_file+".xlsx"
        workbook = load_workbook(excel_file)
        exceldata = []
        grnas = []
        sheet = workbook.active
        #making 
        for value in sheet.iter_rows(values_only=True):
            value = str(value)
            value = value.replace("'", "")
            value = value.replace(",", "")
            value = value.replace("(", "")
            value = value.replace(")", "")
            exceldata.append(value)
        for input in exceldata:
            valid_dna = all(i in cn.VALID_DNA for i in input)
        if valid_dna == True:
            grnas.append(input)
        return grnas

    def get_reverse_complement(dna):
        """
        Converts a 5' to 3' DNA string into its 3' to 5' reverse complement
        Args:
            string: DNA to be converted
        Returns:
            string: reverse complement of input DNA
        """
        dna_comp = [cn.BASE_PAIRS[base] for base in dna]
        dna_rev_comp = dna_comp[::-1]
        return ''.join(dna_rev_comp)

def check_grna(excel_file):
  """
  Takes gRNAs listed in an excel file and checks them for self-targeting or length errors
  Args:
    string: name of excel file listing gRNAs to be checked
  Returns:
    file: excel file listing processed gRNAs and errors if found
  """
  grnas = extractExcelData(excel_file)
  new_grnas = []
  #creating a new blank workbook object
  exceloutput = Workbook()
  #creating a new sheet
  s1 = exceloutput.create_sheet("gRNAcheck")
  if 'Sheet' in exceloutput.sheetnames:
    exceloutput.remove(exceloutput['Sheet'])
  #creating headers
  s1.cell(row = 1, column = 1).value = "gRNAs"
  s1.cell(row = 1, column = 2).value = "self-target error (TTC cut site within gRNA)"
  s1.cell(row = 1, column = 3).value = "length error (>24 nucleotides)"
  s1.cell(row = 1, column = 4).value = "length error (<20 nucleotides)"
  #starting a count
  cell = 0
  for grna in grnas:
    #removing CRISPR cut site within gRNA if present
    grna = grna.removeprefix("TTC")
    grna = grna.removeprefix("ttc")
    #checking if any TTC left within gRNA
    cell = cell + 1
    if 'ttc' in grna or 'TTC' in grna:
      #print("gRNA " + grna + " has a TTC cut site within the gRNA.")
      s1.cell(row = cell+1, column = 2).value = "X"
      #checking gRNA length
    if len(grna) > 24:
      #print("gRNA" + grna + "may be too long (>24 nucleotides).")
      s1.cell(row = cell+1, column = 3).vale = "X"
    if len(grna) < 20:
      #print("gRNA " + grna + " may be too short (<20 nucleotides).")
      c3 = s1.cell(row = cell+1, column = 4).value = "X"
    new_grnas.append(grna)
    c3 = s1.cell(row = cell+1, column = 1)
    c3.value = grna
  exceloutput.save('grnacheck.xlsx')
  return(new_grnas)
  #putting new grnas/updates into an excel file
  
