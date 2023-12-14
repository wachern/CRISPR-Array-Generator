"""
Test cases for crispr_array_generator
Created: Fall 2023
Author: Willow Chernoske
"""

import unittest
import os
import os.path
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from crispr_array_generator.crisprarraygenerator import Array

class TestArray(unittest.TestCase):

    def setUp(self):
        # Forward and reverse DNA strands
        self.forward = 'atgcgga'
        self.reverse = 'tccgcat'

        # TCC and/or non-nucleotide grnas removed
        self.grnas_input = ['ttcaaaggg', 'cccccc', 'tchtaa']
        self.grnas_output1 = ['aaaggg', 'cccccc']
        self.grnas_output2 = ['ttcaaaggg', 'cccccc']
        self.array_output = 'CCCTAAATAATTTCTACTGTTGTAGATaaagggTGGCAAATAATTTCTACTGTTGTAGATcccccc'

    def test_get_reverse_complement(self):
        # Function outputs valid DNA reverse compliment
        result = Array.get_reverse_complement(self.forward)
        self.assertIsNotNone(result)
        self.assertEqual(result, self.reverse)

    def test_extract_excel_data(self):
        # Function can access and correctly process DNA data from an excel file
        path = os.path.join(os.getcwd(), "testfile.xlsx")
        self.assertTrue(os.path.isfile(path))
        result = Array.extract_excel_data('testfile')
        self.assertEqual(result, self.grnas_output2)

    def test_check_grna_array_input(self):
        # Function removes TTC PAM sequence and non-nucleotide characters
        result = Array.check_grna(self.grnas_input)
        self.assertIsNotNone(result)
        self.assertEqual(result, self.grnas_output1)

        # Function creates an excel file "array_report.xlsx"
        path = os.path.join(os.getcwd(), "array_report.xlsx")
        self.assertTrue(os.path.isfile(path))
        if os.path.isfile(path):
            os.remove(path)

    def test_check_grna_excel_input(self):
        # Function removes TTC PAM sequence and non-nucleotide characters
        result = Array.check_grna('testfile')
        self.assertIsNotNone(result)
        self.assertEqual(result, self.grnas_output1)

        # Function creates an excel file "array_report.xlsx"
        path = os.path.join(os.getcwd(), "array_report.xlsx")
        self.assertTrue(os.path.isfile(path))

        # Function correctly identifies length errors
        output = load_workbook("array_report.xlsx")
        sheet = output.active
        result = sheet.cell(row=2 , column=4).value
        self.assertEqual(result, 'X')
        if os.path.isfile(path):
            os.remove(path)

    def test_get_array(self):
        # Function creates an excel file "array_report.xlsx"
        Array.get_array('testfile')
        path = os.path.join(os.getcwd(), "array_report.xlsx")
        self.assertTrue(os.path.isfile(path))

        # Function imports correct array elements into output file
        output = load_workbook("array_report.xlsx")
        sheet = output['Array']
        result = sheet.cell(row=2 , column=2).value
        self.assertEqual(result, self.array_output)
        if os.path.isfile(path):
            os.remove(path)
        # The fact that "array_report" was saved with the "Array" sheet verifies
        # that the make_columns_best_fit function is working since it is
        # required to save the file
